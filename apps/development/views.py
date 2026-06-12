from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Town, Block, Plot, TownMap, PlotCoordinate, TownPartner, PartnerTransaction
import json
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.core.files.storage import default_storage



# ======================================================
# TOWNS
# ======================================================

@login_required
def town_list(request):
    towns = Town.objects.filter(is_deleted=False).order_by('-created_at')
    return render(request, 'development/town_list.html', {'towns': towns})


@login_required
def town_add(request):
    from apps.land.models import LandContract
    contracts = LandContract.objects.filter(
        is_deleted=False,
        status='ACTIVE',
    ).select_related('landlord')

    if request.method == 'POST':
        contract_pk   = request.POST.get('land_contract')
        land_contract = None
        if contract_pk:
            land_contract = LandContract.objects.filter(
                pk=contract_pk,
                is_deleted=False,
                status='ACTIVE',
            ).first()
            if not land_contract:
                messages.error(request, 'Selected land contract is invalid.')
                return render(request, 'development/town_form.html', {
                    'contracts': contracts,
                    'title':     'Add Town',
                })

        town = Town.objects.create(
            name          = request.POST.get('name'),
            name_urdu     = request.POST.get('name_urdu', ''),
            land_contract = land_contract,
            location      = request.POST.get('location', ''),
            total_area    = request.POST.get('total_area') or None,
            description   = request.POST.get('description', ''),
            created_by    = request.user,
        )

        town_image = request.FILES.get('town_image')
        if town_image:
            from apps.development.models import TownMap
            TownMap.objects.create(
                town          = town,
                title         = f"{town.name} — Main Image",
                map_type      = 'IMAGE',
                map_file      = town_image,
                display_order = 0,
                is_active     = True,
                created_by    = request.user,
            )

        messages.success(request, 'Town added successfully.')
        return redirect('development:town_list')

    return render(request, 'development/town_form.html', {
        'contracts': contracts,
        'title':     'Add Town',
    })


@login_required
def town_edit(request, pk):
    from apps.land.models import LandContract
    town      = get_object_or_404(Town, pk=pk, is_deleted=False)
    contracts = LandContract.objects.filter(
        is_deleted=False,
        status='ACTIVE',
    ).select_related('landlord')

    existing_image = town.maps.filter(
        map_type='IMAGE',
        is_active=True
    ).first()

    if request.method == 'POST':
        contract_pk   = request.POST.get('land_contract')
        land_contract = None
        if contract_pk:
            land_contract = LandContract.objects.filter(
                pk=contract_pk,
                is_deleted=False,
                status='ACTIVE',
            ).first()
            if not land_contract:
                messages.error(request, 'Selected land contract is invalid.')
                return render(request, 'development/town_form.html', {
                    'town':           town,
                    'contracts':      contracts,
                    'existing_image': existing_image,
                    'title':          'Edit Town',
                })

        town.name          = request.POST.get('name')
        town.name_urdu     = request.POST.get('name_urdu', '')
        town.land_contract = land_contract
        town.location      = request.POST.get('location', '')
        town.total_area    = request.POST.get('total_area') or None
        town.description   = request.POST.get('description', '')
        town.save()

        town_image = request.FILES.get('town_image')
        if town_image:
            from apps.development.models import TownMap
            if existing_image:
                existing_image.map_file = town_image
                existing_image.save()
            else:
                TownMap.objects.create(
                    town          = town,
                    title         = f"{town.name} — Main Image",
                    map_type      = 'IMAGE',
                    map_file      = town_image,
                    display_order = 0,
                    is_active     = True,
                    created_by    = request.user,
                )

        messages.success(request, 'Town updated.')
        return redirect('development:town_list')

    return render(request, 'development/town_form.html', {
        'town':           town,
        'contracts':      contracts,
        'existing_image': existing_image,
        'title':          'Edit Town',
    })


@login_required
def town_delete(request, pk):
    town            = get_object_or_404(Town, pk=pk)
    town.is_deleted = True
    town.save()
    messages.success(request, 'Town deleted.')
    return redirect('development:town_list')


# ======================================================
# TOWN DETAIL / PARTNERS
# ======================================================

@login_required
def town_detail(request, pk):
    town     = get_object_or_404(Town, pk=pk, is_deleted=False)
    partners = TownPartner.objects.filter(town=town, is_deleted=False).prefetch_related('transactions')
    return render(request, 'development/town_detail.html', {
        'town':     town,
        'partners': partners,
    })


@login_required
def partner_add(request, town_pk):
    town = get_object_or_404(Town, pk=town_pk, is_deleted=False)
    if request.method == 'POST':
        TownPartner.objects.create(
            town          = town,
            name          = request.POST.get('name', '').strip(),
            share_percent = request.POST.get('share_percent', 0),
            notes         = request.POST.get('notes', ''),
            created_by    = request.user,
        )
        messages.success(request, 'Partner added.')
    return redirect('development:town_detail', pk=town_pk)


@login_required
def partner_edit(request, pk):
    partner = get_object_or_404(TownPartner, pk=pk, is_deleted=False)
    if request.method == 'POST':
        partner.name          = request.POST.get('name', '').strip()
        partner.share_percent = request.POST.get('share_percent', 0)
        partner.notes         = request.POST.get('notes', '')
        partner.save()
        messages.success(request, 'Partner updated.')
    return redirect('development:town_detail', pk=partner.town_id)


@login_required
def partner_delete(request, pk):
    partner            = get_object_or_404(TownPartner, pk=pk)
    partner.is_deleted = True
    partner.save()
    messages.success(request, 'Partner removed.')
    return redirect('development:town_detail', pk=partner.town_id)


@login_required
def transaction_add(request, partner_pk):
    partner = get_object_or_404(TownPartner, pk=partner_pk, is_deleted=False)
    if request.method == 'POST':
        PartnerTransaction.objects.create(
            partner          = partner,
            transaction_type = request.POST.get('transaction_type'),
            amount           = request.POST.get('amount'),
            transaction_date = request.POST.get('transaction_date'),
            narration        = request.POST.get('narration', ''),
            created_by       = request.user,
        )
        messages.success(request, 'Transaction recorded.')
    return redirect('development:town_detail', pk=partner.town_id)


@login_required
def transaction_delete(request, pk):
    txn = get_object_or_404(PartnerTransaction, pk=pk)
    town_pk = txn.partner.town_id
    txn.delete()
    messages.success(request, 'Transaction deleted.')
    return redirect('development:town_detail', pk=town_pk)


# ======================================================
# BLOCKS
# ======================================================

@login_required
def block_list(request):
    town_id = request.GET.get('town')
    towns   = Town.objects.filter(is_deleted=False).order_by('name')
    blocks  = Block.objects.filter(is_deleted=False).order_by('town', 'name')

    if town_id:
        blocks = blocks.filter(town_id=town_id)

    return render(request, 'development/block_list.html', {
        'blocks':        blocks,
        'towns':         towns,
        'selected_town': town_id,
    })


@login_required
def block_add(request):
    towns = Town.objects.filter(is_deleted=False).order_by('name')

    if request.method == 'POST':
        town = get_object_or_404(Town, pk=request.POST.get('town'))
        Block.objects.create(
            town        = town,
            name        = request.POST.get('name'),
            name_urdu   = request.POST.get('name_urdu', ''),
            total_plots = request.POST.get('total_plots', 0),
            description = request.POST.get('description', ''),
            created_by  = request.user,
        )
        messages.success(request, 'Block added successfully.')
        return redirect('development:block_list')

    return render(request, 'development/block_form.html', {
        'towns': towns,
        'title': 'Add Block',
    })


@login_required
def block_edit(request, pk):
    block = get_object_or_404(Block, pk=pk, is_deleted=False)
    towns = Town.objects.filter(is_deleted=False).order_by('name')

    if request.method == 'POST':
        block.town        = get_object_or_404(Town, pk=request.POST.get('town'))
        block.name        = request.POST.get('name')
        block.name_urdu   = request.POST.get('name_urdu', '')
        block.total_plots = request.POST.get('total_plots', 0)
        block.description = request.POST.get('description', '')
        block.save()
        messages.success(request, 'Block updated.')
        return redirect('development:block_list')

    return render(request, 'development/block_form.html', {
        'block': block,
        'towns': towns,
        'title': 'Edit Block',
    })


@login_required
def block_delete(request, pk):
    block            = get_object_or_404(Block, pk=pk)
    block.is_deleted = True
    block.save()
    messages.success(request, 'Block deleted.')
    return redirect('development:block_list')


# ======================================================
# PLOTS
# ======================================================

@login_required
def plot_list(request):
    block_id = request.GET.get('block')
    town_id  = request.GET.get('town')
    status   = request.GET.get('status')

    towns      = Town.objects.filter(is_deleted=False).order_by('name')
    blocks     = Block.objects.filter(is_deleted=False).order_by('town', 'name')
    plots = Plot.objects.filter(is_deleted=False).select_related('block__town').extra(
    select={
        'plot_alpha': "TRIM(TRIM(plot_no, '0123456789'), ' ')",
        'plot_num':   "CAST(COALESCE(NULLIF(TRIM(plot_no, 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz- '), ''), '0') AS INTEGER)",
    }
        )
 
    base_plots = Plot.objects.filter(is_deleted=False)

    if town_id:
        plots      = plots.filter(block__town_id=town_id)
        blocks     = blocks.filter(town_id=town_id)
        base_plots = base_plots.filter(block__town_id=town_id)
    if block_id:
        plots      = plots.filter(block_id=block_id)
        base_plots = base_plots.filter(block_id=block_id)
    if status:
        plots = plots.filter(status=status)

    plots = plots.order_by('block__town__name', 'block__name', 'plot_alpha', 'plot_num')

    status_counts = {}
    for val, label in Plot.STATUS_CHOICES:
        status_counts[val] = base_plots.filter(status=val).count()

    # Attach booking financials to each plot
    from apps.sales.models import Booking
    from django.db.models import Sum as _Sum
    booking_map = {
        b.plot_id: b
        for b in Booking.objects
            .filter(is_deleted=False, status__in=['ACTIVE', 'COMPLETED'])
            .annotate(amount_received=_Sum('receipts__amount'))
    }

    def to_marlas(size, unit):
        size = float(size or 0)
        if unit == 'KANAL':
            return size * 20
        elif unit == 'SQFT':
            return size / 270
        return size  # MARLA

    # Group by town
    from collections import OrderedDict
    towns_data = OrderedDict()

    for p in plots:
        town_name = p.block.town.name
        if town_name not in towns_data:
            towns_data[town_name] = []
        bk            = booking_map.get(p.pk)
        display_price = (bk.net_price or p.price or 0) if bk else (p.price or 0)
        received      = (bk.amount_received or 0) if bk else 0
        balance       = display_price - received
        marlas        = to_marlas(p.size, p.size_unit)
        towns_data[town_name].append({
            'obj':           p,
            'display_price': display_price,
            'received':      received,
            'balance':       balance,
            'marlas':        marlas,
        })

    # Build grouped list with per-town totals
    grouped = []
    grand_price = grand_received = grand_balance = grand_marlas = 0

    for town_name, items in towns_data.items():
        t_price    = sum(i['display_price'] for i in items)
        t_received = sum(i['received']      for i in items)
        t_balance  = sum(i['balance']       for i in items)
        t_marlas   = sum(i['marlas']        for i in items)
        grand_price    += t_price
        grand_received += t_received
        grand_balance  += t_balance
        grand_marlas   += t_marlas
        grouped.append({
            'town_name':    town_name,
            'items':        items,
            't_price':      t_price,
            't_received':   t_received,
            't_balance':    t_balance,
            't_marlas':     t_marlas,
        })

    # flat list still needed for status badges
    plots_data = [i for g in grouped for i in g['items']]
    total_price    = grand_price
    total_received = grand_received
    total_balance  = grand_balance

    return render(request, 'development/plot_list.html', {
        'plots':           plots_data,
        'grouped':         grouped,
        'towns':           towns,
        'blocks':          blocks,
        'selected_town':   town_id,
        'selected_block':  block_id,
        'selected_status': status,
        'status_choices':  Plot.STATUS_CHOICES,
        'status_counts':   status_counts,
        'total_price':     total_price,
        'total_received':  total_received,
        'total_balance':   total_balance,
        'grand_marlas':    grand_marlas,
    })


@login_required
def plot_add(request):
    blocks = Block.objects.filter(
        is_deleted=False
    ).select_related('town').order_by('town__name', 'name')

    towns = Town.objects.filter(is_deleted=False).order_by('name')

    selected_block_id = request.GET.get('block', '')
    selected_town_id  = request.GET.get('town', '')

    existing_plots = []
    if selected_block_id:
        existing_plots = Plot.objects.filter(
            block_id=selected_block_id,
            is_deleted=False
        ).extra(
            select={
                'plot_alpha': "TRIM(TRIM(plot_no, '0123456789'), ' ')",
                'plot_num':   "CAST(COALESCE(NULLIF(TRIM(plot_no, 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz- '), ''), '0') AS INTEGER)",
            }
        ).order_by('plot_alpha', 'plot_num')
    elif selected_town_id:
        existing_plots = Plot.objects.filter(
            block__town_id=selected_town_id,
            is_deleted=False
        ).select_related('block').extra(
            select={
                'plot_alpha': "TRIM(TRIM(plot_no, '0123456789'), ' ')",
                'plot_num':   "CAST(COALESCE(NULLIF(TRIM(plot_no, 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz- '), ''), '0') AS INTEGER)",
            }
        ).order_by('block__name', 'plot_alpha', 'plot_num')

    if request.method == 'POST':
        block        = get_object_or_404(Block, pk=request.POST.get('block'))
        size_marla   = float(request.POST.get('size_marla') or 0)
        size_sqft    = float(request.POST.get('size_sqft')  or 0)
        total_marlas = size_marla + size_sqft / 270
        Plot.objects.create(
            block      = block,
            plot_no    = request.POST.get('plot_no'),
            size       = round(total_marlas, 4),
            size_unit  = 'MARLA',
            plot_type  = request.POST.get('plot_type', 'RESIDENTIAL'),
            price      = request.POST.get('price'),
            status     = request.POST.get('status', 'AVAILABLE'),
            notes      = request.POST.get('notes', ''),
            created_by = request.user,
        )
        messages.success(request, 'Plot added successfully.')
        return redirect(request.get_full_path())

    return render(request, 'development/plot_form.html', {
        'blocks':             blocks,
        'towns':              towns,
        'title':              'Add Plot',
        'existing_plots':     existing_plots,
        'selected_block_id':  selected_block_id,
        'selected_town_id':   selected_town_id,
        'plot_marla':         '',
        'plot_sqft':          '',
    })

@login_required
def plot_edit(request, pk):
    plot   = get_object_or_404(Plot, pk=pk, is_deleted=False)
    blocks = Block.objects.filter(
        is_deleted=False
    ).select_related('town').order_by('town__name', 'name')

    if request.method == 'POST':
        size_marla   = float(request.POST.get('size_marla') or 0)
        size_sqft    = float(request.POST.get('size_sqft')  or 0)
        total_marlas = size_marla + size_sqft / 270
        plot.block     = get_object_or_404(Block, pk=request.POST.get('block'))
        plot.plot_no   = request.POST.get('plot_no')
        plot.size      = round(total_marlas, 4)
        plot.size_unit = 'MARLA'
        plot.plot_type = request.POST.get('plot_type', 'RESIDENTIAL')
        plot.price     = request.POST.get('price')
        plot.status    = request.POST.get('status', 'AVAILABLE')
        plot.notes     = request.POST.get('notes', '')
        plot.save()
        messages.success(request, 'Plot updated.')
        return redirect('development:plot_list')

    # Convert existing size to marla + sqft for pre-filling the form
    existing_marlas = float(plot.size or 0)
    if plot.size_unit == 'KANAL':
        existing_marlas *= 20
    elif plot.size_unit == 'SQFT':
        existing_marlas /= 270
    plot_marla = int(existing_marlas)
    plot_sqft  = round((existing_marlas - plot_marla) * 270)

    from apps.sales.models import Booking as _Booking
    active_booking = _Booking.objects.filter(
        plot=plot, is_deleted=False, status__in=['ACTIVE', 'COMPLETED']
    ).first()

    return render(request, 'development/plot_form.html', {
        'plot':           plot,
        'blocks':         blocks,
        'title':          'Edit Plot',
        'plot_marla':     plot_marla,
        'plot_sqft':      plot_sqft,
        'active_booking': active_booking,
    })


@login_required
def plot_delete(request, pk):
    plot = get_object_or_404(Plot, pk=pk, is_deleted=False)

    if plot.status in ('BOOKED', 'SOLD'):
        messages.error(request, 'Cannot delete a booked or sold plot.')
        return redirect('development:plot_list')

    plot.is_deleted = True
    plot.save()
    messages.success(request, 'Plot deleted.')
    return redirect('development:plot_list')


# ======================================================
# TOWN MAP VIEWS
# ======================================================

@login_required
def map_list(request, town_pk):
    town = get_object_or_404(Town, pk=town_pk)
    maps = TownMap.objects.filter(town=town).select_related('block')
    return render(request, 'development/map_list.html', {
        'town': town,
        'maps': maps,
    })


@login_required
def map_upload(request, town_pk):
    town   = get_object_or_404(Town, pk=town_pk)
    blocks = Block.objects.filter(town=town, is_deleted=False).order_by('name')

    if request.method == 'POST':
        title         = request.POST.get('title', '').strip()
        map_type      = request.POST.get('map_type', 'IMAGE')
        block_id      = request.POST.get('block', '')
        display_order = request.POST.get('display_order', 0)
        map_file      = request.FILES.get('map_file')

        if not title:
            messages.error(request, 'Title is required.')
            return render(request, 'development/map_upload.html', {
                'town': town, 'blocks': blocks
            })

        if not map_file:
            messages.error(request, 'Please select a file to upload.')
            return render(request, 'development/map_upload.html', {
                'town': town, 'blocks': blocks
            })

        fname = map_file.name.lower()
        if map_type == 'IMAGE' and not (
            fname.endswith('.jpg') or
            fname.endswith('.jpeg') or
            fname.endswith('.png')
        ):
            messages.error(request, 'Image maps must be JPG or PNG.')
            return render(request, 'development/map_upload.html', {
                'town': town, 'blocks': blocks
            })

        if map_type == 'PDF' and not fname.endswith('.pdf'):
            messages.error(request, 'PDF maps must be .pdf files.')
            return render(request, 'development/map_upload.html', {
                'town': town, 'blocks': blocks
            })

        block = Block.objects.filter(pk=block_id).first() if block_id else None

        TownMap.objects.create(
            town          = town,
            block         = block,
            title         = title,
            map_type      = map_type,
            map_file      = map_file,
            display_order = display_order or 0,
            is_active     = True,
            created_by    = request.user,
        )

        messages.success(request, f'Map "{title}" uploaded successfully.')
        return redirect('development:map_list', town_pk=town.pk)

    return render(request, 'development/map_upload.html', {
        'town':   town,
        'blocks': blocks,
    })


@login_required
def map_edit(request, pk):
    town_map = get_object_or_404(TownMap, pk=pk)
    town     = town_map.town
    blocks   = Block.objects.filter(town=town, is_deleted=False).order_by('name')

    if request.method == 'POST':
        town_map.title         = request.POST.get('title', town_map.title).strip()
        town_map.display_order = request.POST.get('display_order', 0)
        town_map.is_active     = request.POST.get('is_active') == 'on'
        block_id               = request.POST.get('block', '')
        town_map.block         = Block.objects.filter(pk=block_id).first() if block_id else None

        new_file = request.FILES.get('map_file')
        if new_file:
            town_map.map_file = new_file

        town_map.save()
        messages.success(request, 'Map updated successfully.')
        return redirect('development:map_list', town_pk=town.pk)

    return render(request, 'development/map_edit.html', {
        'town_map': town_map,
        'town':     town,
        'blocks':   blocks,
    })


@login_required
def map_delete(request, pk):
    town_map = get_object_or_404(TownMap, pk=pk)
    town_pk  = town_map.town.pk

    if request.method == 'POST':
        town_map.map_file.delete(save=False)
        town_map.delete()
        messages.success(request, 'Map deleted.')
        return redirect('development:map_list', town_pk=town_pk)

    return render(request, 'development/map_confirm_delete.html', {
        'town_map': town_map,
    })


@login_required
def map_toggle(request, pk):
    town_map           = get_object_or_404(TownMap, pk=pk)
    town_map.is_active = not town_map.is_active
    town_map.save()
    status = 'activated' if town_map.is_active else 'deactivated'
    messages.success(request, f'Map {status}.')
    return redirect('development:map_list', town_pk=town_map.town.pk)


# ======================================================
# MAP EDITOR (SUPERUSER ONLY)
# ======================================================

@login_required
def map_editor(request, pk):
    if request.user.role != 'SUPERUSER':
        messages.error(request, 'Access denied. Superuser only.')
        return redirect('development:map_list',
                        town_pk=TownMap.objects.get(pk=pk).town.pk)

    town_map = get_object_or_404(TownMap, pk=pk)
    town     = town_map.town

    plots = Plot.objects.filter(
        block__town=town,
        is_deleted=False
    ).select_related('block').extra(
        select={
            'plot_alpha': "TRIM(TRIM(plot_no, '0123456789'), ' ')",
            'plot_num':   "CAST(COALESCE(NULLIF(TRIM(plot_no, 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz- '), ''), '0') AS INTEGER)",
        }
    ).order_by('block__name', 'plot_alpha', 'plot_num')

    existing = PlotCoordinate.objects.filter(
        town_map=town_map
    ).select_related('plot')

    coords_data = {}
    for c in existing:
        coords_data[str(c.plot.pk)] = {
            'id':          c.pk,
            'plot_id':     c.plot.pk,
            'plot_no':     c.plot.plot_no,
            'plot_status': c.plot.status,
            'coordinates': c.coordinates,
        }

    marked_plot_ids = [c.plot.pk for c in existing]

    return render(request, 'development/map_editor.html', {
        'town_map':        town_map,
        'town':            town,
        'plots':           plots,
        'coords_json':     json.dumps(coords_data),
        'marked_plot_ids': marked_plot_ids,
    })


@login_required
@require_POST
def map_save_coordinate(request, map_pk):
    if request.user.role != 'SUPERUSER':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        data        = json.loads(request.body)
        plot_id     = data.get('plot_id')
        coordinates = data.get('coordinates')

        if not plot_id or not coordinates:
            return JsonResponse(
                {'error': 'plot_id and coordinates required'}, status=400)

        town_map = get_object_or_404(TownMap, pk=map_pk)
        plot     = get_object_or_404(Plot, pk=plot_id)

        coord, created = PlotCoordinate.objects.update_or_create(
            town_map = town_map,
            plot     = plot,
            defaults = {
                'coordinates': coordinates,
                'created_by':  request.user,
            }
        )

        return JsonResponse({
            'success': True,
            'id':      coord.pk,
            'plot_no': plot.plot_no,
            'created': created,
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def map_delete_coordinate(request, coord_pk):
    if request.user.role != 'SUPERUSER':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        coord   = get_object_or_404(PlotCoordinate, pk=coord_pk)
        plot_no = coord.plot.plot_no
        coord.delete()
        return JsonResponse({'success': True, 'plot_no': plot_no})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def map_get_coordinates(request, map_pk):
    town_map = get_object_or_404(TownMap, pk=map_pk)
    existing = PlotCoordinate.objects.filter(
        town_map=town_map
    ).select_related('plot')

    coords_data = []
    for c in existing:
        coords_data.append({
            'id':          c.pk,
            'plot_id':     c.plot.pk,
            'plot_no':     c.plot.plot_no,
            'plot_status': c.plot.status,
            'size':        str(c.plot.size),
            'size_unit':   c.plot.size_unit,
            'plot_type':   c.plot.type,
            'price':       str(c.plot.price),
            'coordinates': c.coordinates,
        })

    return JsonResponse({'coordinates': coords_data})


# ======================================================
# MAP VIEWER
# ======================================================

@login_required
def map_viewer(request, pk):
    town_map = get_object_or_404(TownMap, pk=pk)
    return render(request, 'development/map_viewer.html', {
        'town_map': town_map,
    })


# ======================================================
# PDF MAP VIEWER
# ======================================================

@login_required
def map_pdf_viewer(request, pk):
    town_map = get_object_or_404(TownMap, pk=pk)
    if not town_map.is_pdf:
        messages.error(request, 'This map is not a PDF.')
        return redirect('development:map_viewer', pk=pk)
    return render(request, 'development/map_pdf_viewer.html', {
        'town_map': town_map,
    })


@login_required
def plot_bulk_add(request):
    blocks = Block.objects.filter(
        is_deleted=False
    ).select_related('town').order_by('town__name', 'name')

    towns = Town.objects.filter(is_deleted=False).order_by('name')

    selected_block_id = request.GET.get('block', '')
    selected_town_id  = request.GET.get('town', '')

    existing_plots = []
    if selected_block_id:
        existing_plots = Plot.objects.filter(
            block_id=selected_block_id,
            is_deleted=False
        ).extra(
            select={
                'plot_alpha': "TRIM(TRIM(plot_no, '0123456789'), ' ')",
                'plot_num':   "CAST(COALESCE(NULLIF(TRIM(plot_no, 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz- '), ''), '0') AS INTEGER)",
            }
        ).order_by('plot_alpha', 'plot_num')
    elif selected_town_id:
        existing_plots = Plot.objects.filter(
            block__town_id=selected_town_id,
            is_deleted=False
        ).select_related('block').extra(
            select={
                'plot_alpha': "TRIM(TRIM(plot_no, '0123456789'), ' ')",
                'plot_num':   "CAST(COALESCE(NULLIF(TRIM(plot_no, 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz- '), ''), '0') AS INTEGER)",
            }
        ).order_by('block__name', 'plot_alpha', 'plot_num')

    if request.method == 'POST':
        import json
        plots_data = request.POST.get('plots_json', '[]')
        try:
            plots_list = json.loads(plots_data)
        except Exception:
            messages.error(request, 'Invalid data submitted.')
            return redirect('development:plot_bulk_add')

        created = 0
        errors  = []
        for i, p in enumerate(plots_list):
            try:
                block = Block.objects.get(pk=p['block'])
                Plot.objects.create(
                    block      = block,
                    plot_no    = p['plot_no'],
                    size       = p['size'],
                    size_unit  = p.get('size_unit', 'MARLA'),
                    plot_type  = p.get('plot_type', 'RESIDENTIAL'),
                    price      = p['price'],
                    status     = p.get('status', 'AVAILABLE'),
                    notes      = p.get('notes', ''),
                    created_by = request.user,
                )
                created += 1
            except Exception as e:
                errors.append(f"Row {i+1} ({p.get('plot_no','?')}): {str(e)}")

        if errors:
            for err in errors:
                messages.error(request, err)
        if created:
            messages.success(request, f'{created} plot(s) saved successfully.')
        return redirect(request.get_full_path())

    return render(request, 'development/plot_bulk_add.html', {
        'blocks':            blocks,
        'towns':             towns,
        'title':             'Bulk Add Plots',
        'existing_plots':    existing_plots,
        'selected_block_id': selected_block_id,
        'selected_town_id':  selected_town_id,
    })

@login_required
def check_plot_availability(request):
    plot_no  = request.GET.get('plot_no', '').strip()
    block_id = request.GET.get('block_id', '').strip()
    plot_pk  = request.GET.get('plot_pk', '').strip()

    if not plot_no or not block_id:
        return JsonResponse({'available': True, 'plot': None})

    qs = Plot.objects.filter(
        plot_no__iexact=plot_no,
        block_id=block_id,
        is_deleted=False,
    )

    if plot_pk:
        qs = qs.exclude(pk=plot_pk)

    existing = qs.select_related('block__town').first()

    if existing:
        return JsonResponse({
            'available': False,
            'plot': {
                'plot_no':   existing.plot_no,
                'block':     existing.block.name,
                'town':      existing.block.town.name,
                'size':      str(existing.size),
                'size_unit': existing.size_unit,
                'plot_type': existing.get_plot_type_display(),
                'price':     str(existing.price),
                'status':    existing.status,
            }
        })

    return JsonResponse({'available': True, 'plot': None})


@login_required
def check_town_name(request):
    name    = request.GET.get('name', '').strip()
    town_pk = request.GET.get('town_pk', '').strip()

    if not name:
        return JsonResponse({'available': True, 'town': None})

    qs = Town.objects.filter(name__iexact=name, is_deleted=False)

    if town_pk:
        qs = qs.exclude(pk=town_pk)

    existing = qs.first()

    if existing:
        return JsonResponse({
            'available': False,
            'town': {
                'name':     existing.name,
                'location': existing.location or '—',
                'blocks':   existing.blocks.filter(is_deleted=False).count(),
            }
        })

    return JsonResponse({'available': True, 'town': None})


@login_required
def check_block_name(request):
    name     = request.GET.get('name', '').strip()
    town_id  = request.GET.get('town_id', '').strip()
    block_pk = request.GET.get('block_pk', '').strip()

    if not name or not town_id:
        return JsonResponse({'available': True, 'block': None})

    qs = Block.objects.filter(
        name__iexact=name,
        town_id=town_id,
        is_deleted=False
    )

    if block_pk:
        qs = qs.exclude(pk=block_pk)

    existing = qs.select_related('town').first()

    if existing:
        return JsonResponse({
            'available': False,
            'block': {
                'name':        existing.name,
                'town':        existing.town.name,
                'total_plots': existing.total_plots,
                'available':   existing.available_plots,
                'sold':        existing.sold_plots,
            }
        })

    return JsonResponse({'available': True, 'block': None})


# ======================================================
# PLOTS — EXCEL EXPORT
# ======================================================

def _get_plots_queryset(request):
    """Shared filter logic for plot exports."""
    block_id = request.GET.get('block', '').strip()
    town_id  = request.GET.get('town', '').strip()
    status   = request.GET.get('status', '').strip()

    # Treat 'None' string or empty as no filter
    if block_id in ('', 'None'):
        block_id = None
    if town_id in ('', 'None'):
        town_id = None
    if status in ('', 'None'):
        status = None

    plots = Plot.objects.filter(is_deleted=False).select_related('block__town').extra(
        select={
            'plot_alpha': "TRIM(TRIM(plot_no, '0123456789'), ' ')",
            'plot_num':   "CAST(COALESCE(NULLIF(TRIM(plot_no, 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz- '), ''), '0') AS INTEGER)",
        }
    )
    if town_id:
        plots = plots.filter(block__town_id=town_id)
    if block_id:
        plots = plots.filter(block_id=block_id)
    if status:
        plots = plots.filter(status=status)
    return plots.order_by('block__town__name', 'block__name', 'plot_alpha', 'plot_num')


def _to_marlas(size, unit):
    size = float(size or 0)
    if unit == 'KANAL':
        return size * 20
    elif unit == 'SQFT':
        return size / 270
    return size


@login_required
def plot_list_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from apps.sales.models import Booking
    from django.db.models import Sum as _Sum

    plots = _get_plots_queryset(request)
    booking_map = {
        b.plot_id: b
        for b in Booking.objects
            .filter(is_deleted=False, status__in=['ACTIVE', 'COMPLETED'])
            .annotate(amount_received=_Sum('receipts__amount'))
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Plots'

    # Styles
    hdr_font   = Font(bold=True, color='FFFFFF')
    hdr_fill   = PatternFill('solid', fgColor='1F3864')
    town_font  = Font(bold=True, color='FFFFFF')
    town_fill  = PatternFill('solid', fgColor='2E75B6')
    sub_fill   = PatternFill('solid', fgColor='D6E4F0')
    total_font = Font(bold=True)
    total_fill = PatternFill('solid', fgColor='BDD7EE')
    grand_fill = PatternFill('solid', fgColor='1F3864')
    grand_font = Font(bold=True, color='FFFFFF')
    center     = Alignment(horizontal='center', vertical='center')
    right      = Alignment(horizontal='right')
    thin       = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    cols = ['#', 'Plot No', 'Block', 'Size', 'Unit', 'Marlas', 'Type',
            'Price (Rs.)', 'Received (Rs.)', 'Balance (Rs.)', 'Status']
    ws.append(cols)
    for c in range(1, len(cols)+1):
        cell = ws.cell(row=1, column=c)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = thin

    from collections import OrderedDict
    towns_data = OrderedDict()
    for p in plots:
        tn = p.block.town.name
        if tn not in towns_data:
            towns_data[tn] = []
        bk            = booking_map.get(p.pk)
        display_price = (bk.net_price or p.price or 0) if bk else (p.price or 0)
        received      = (bk.amount_received or 0) if bk else 0
        balance       = float(display_price) - float(received)
        marlas        = _to_marlas(p.size, p.size_unit)
        towns_data[tn].append((p, display_price, received, balance, marlas))

    row_num = 2
    grand_price = grand_received = grand_balance = grand_marlas = 0

    for town_name, items in towns_data.items():
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=len(cols))
        tc = ws.cell(row=row_num, column=1, value=f'  {town_name}')
        tc.font = town_font; tc.fill = town_fill
        tc.alignment = Alignment(horizontal='left', vertical='center')
        tc.border = thin
        row_num += 1

        t_price = t_received = t_balance = t_marlas = 0
        for idx, (p, dp, rec, bal, mar) in enumerate(items, 1):
            ws.append([idx, p.plot_no, p.block.name,
                       float(p.size), p.size_unit, round(mar, 2),
                       p.plot_type, float(dp), float(rec), round(bal, 2),
                       p.status])
            for c in range(1, len(cols)+1):
                cell = ws.cell(row=row_num, column=c)
                cell.border = thin
                if c in (8, 9, 10):
                    cell.alignment = right
                if idx % 2 == 0:
                    cell.fill = sub_fill
            row_num += 1
            t_price    += float(dp)
            t_received += float(rec)
            t_balance  += bal
            t_marlas   += mar

        ws.append(['', f'Total — {town_name}', '', '', '', round(t_marlas, 2),
                   '', t_price, t_received, round(t_balance, 2), ''])
        for c in range(1, len(cols)+1):
            cell = ws.cell(row=row_num, column=c)
            cell.font = total_font; cell.fill = total_fill; cell.border = thin
            if c in (8, 9, 10):
                cell.alignment = right
        row_num += 1

        grand_price    += t_price
        grand_received += t_received
        grand_balance  += t_balance
        grand_marlas   += t_marlas

    ws.append(['', 'GRAND TOTAL', '', '', '', round(grand_marlas, 2),
               '', grand_price, grand_received, round(grand_balance, 2), ''])
    for c in range(1, len(cols)+1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = grand_font; cell.fill = grand_fill; cell.border = thin
        if c in (8, 9, 10):
            cell.alignment = right

    widths = [5, 12, 14, 8, 8, 10, 14, 18, 18, 18, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plots.xlsx"'
    wb.save(response)
    return response


# ======================================================
# PLOTS — PRINT VIEW
# ======================================================

@login_required
def plot_list_print(request):
    from apps.sales.models import Booking
    from django.db.models import Sum as _Sum
    from collections import OrderedDict

    plots = _get_plots_queryset(request)
    booking_map = {
        b.plot_id: b
        for b in Booking.objects
            .filter(is_deleted=False, status__in=['ACTIVE', 'COMPLETED'])
            .annotate(amount_received=_Sum('receipts__amount'))
    }

    towns_data = OrderedDict()
    for p in plots:
        tn = p.block.town.name
        if tn not in towns_data:
            towns_data[tn] = []
        bk            = booking_map.get(p.pk)
        display_price = (bk.net_price or p.price or 0) if bk else (p.price or 0)
        received      = (bk.amount_received or 0) if bk else 0
        balance       = float(display_price) - float(received)
        marlas        = _to_marlas(p.size, p.size_unit)
        towns_data[tn].append({
            'obj':           p,
            'display_price': display_price,
            'received':      received,
            'balance':       balance,
            'marlas':        marlas,
        })

    grouped = []
    grand_price = grand_received = grand_balance = grand_marlas = 0
    for town_name, items in towns_data.items():
        t_price    = sum(float(i['display_price']) for i in items)
        t_received = sum(float(i['received'])      for i in items)
        t_balance  = sum(i['balance']              for i in items)
        t_marlas   = sum(i['marlas']               for i in items)
        grand_price    += t_price
        grand_received += t_received
        grand_balance  += t_balance
        grand_marlas   += t_marlas
        grouped.append({
            'town_name':  town_name,
            'items':      items,
            't_price':    t_price,
            't_received': t_received,
            't_balance':  t_balance,
            't_marlas':   t_marlas,
        })

    return render(request, 'development/plot_list_print.html', {
        'grouped':        grouped,
        'grand_price':    grand_price,
        'grand_received': grand_received,
        'grand_balance':  grand_balance,
        'grand_marlas':   grand_marlas,
    })